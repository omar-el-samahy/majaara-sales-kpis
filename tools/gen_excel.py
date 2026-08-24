from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

NAVY = "16324F"
TEAL = "0D9488"
YELLOW = "FFF9C4"
GRAY = "F1F5F9"
WHITE = "FFFFFF"

title_font = Font(bold=True, size=14, color=WHITE)
section_font = Font(bold=True, size=11, color=NAVY)
label_font = Font(size=11)
input_fill = PatternFill("solid", fgColor=YELLOW)
calc_fill = PatternFill("solid", fgColor=GRAY)
header_fill = PatternFill("solid", fgColor=NAVY)
teal_fill = PatternFill("solid", fgColor=TEAL)
white_font = Font(bold=True, color=WHITE, size=11)
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(wrap_text=True, vertical="center")

wb = Workbook()

def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = title_font
    c.fill = header_fill
    c.alignment = center
    for row in ws[cell_range]:
        for x in row:
            x.fill = header_fill

def put(ws, cell, value, fill=None, font=None, fmt=None, align=None):
    c = ws[cell]
    c.value = value
    if fill: c.fill = fill
    if font: c.font = font
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    c.border = border
    return c

ws_i = wb.active
ws_i.title = "Instructions"
style_title(ws_i, "A1:F1", "نظام مؤشرات الأداء — Majaara KPI Scorecards")
lines = [
    "",
    "كيفية الاستخدام | How to use:",
    "1. املأ الخلايا الصفراء فقط (الأهداف والمنجزات) — النقاط تُحتسب تلقائياً.",
    "1. Fill the YELLOW cells only (targets & actuals) — points are calculated automatically.",
    "2. ورقة 'Sales Rep' لمندوبي المبيعات (100 نقطة) وورقة 'Team Leader' لقادة الفرق (110 نقاط).",
    "2. Sheet 'Sales Rep' is for sales representatives (100 pts), sheet 'Team Leader' for team leaders (110 pts).",
    "3. التقييم: ممتاز / جيد جداً / يحتاج تحسين / مراجعة — يظهر تلقائياً مع تنسيق ملون.",
    "3. Rating: Excellent / Very Good / Needs Improvement / Under Review — shown automatically with colors.",
    "",
    "ملاحظة: فئات النقاط مطابقة لوثيقة KPIs الرسمية. بعض فئات المكافأة (تجاوز الهدف) تعطي نقاطاً أعلى من الحد الأقصى الموجز في جدول الملخص، لذا قد يتجاوز المجموع 100/110 بشكل طفيف عند تجاوز الأهداف.",
    "Note: tiers follow the official KPI document verbatim. Bonus tiers may slightly exceed the 100/110 summary max when targets are exceeded.",
]
for i, ln in enumerate(lines, start=2):
    ws_i.merge_cells(f"A{i}:F{i}")
    c = ws_i[f"A{i}"]
    c.value = ln
    c.alignment = wrap
    if ln.endswith(":"):
        c.font = section_font
for col in "ABCDEF":
    ws_i.column_dimensions[col].width = 22

def build_rep_sheet(ws):
    style_title(ws, "A1:E1", "بطاقة تقييم مندوب المبيعات | Sales Representative Scorecard")
    put(ws, "A3", "الأهداف الشهرية | Monthly Targets", font=section_font)
    rows_t = [("هدف المبيعات | Sales target", "B4"), ("هدف الاجتماعات | Meetings target", "B5"), ("دورة البيع المستهدفة (يوم) | Target cycle (days)", "B6")]
    for i, (lbl, cell) in enumerate(rows_t):
        put(ws, f"A{4+i}", lbl, font=label_font, align=wrap)
        put(ws, cell, None, input_fill)

    put(ws, "A8", "المنجز الفعلي | Actuals", font=section_font)
    actuals = [
        ("المبيعات المحققة | Sales achieved", "B9", "#,##0"),
        ("الاجتماعات المنفذة | Meetings held", "B10", "0"),
        ("الصفقات المغلقة | Deals closed", "B11", "0"),
        ("متوسط دورة البيع (يوم) | Avg sales cycle (days)", "B12", "0.0"),
        ("الالتزام بـ CRM (%) | CRM compliance %", "B13", "0%"),
        ("متوسط تقييم العملاء (0-10) | Avg customer rating", "B14", "0.0"),
    ]
    for i, (lbl, cell, fmt) in enumerate(actuals):
        put(ws, f"A{9+i}", lbl, font=label_font, align=wrap)
        put(ws, cell, None, input_fill, fmt=fmt)

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="10", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B14")

    headers = ["المؤشر | KPI", "المنجز | Actual", "الفئة | Tier", "النقاط | Points", "الحد | Max"]
    for j, h in enumerate(headers):
        put(ws, f"{get_column_letter(j+1)}17", h, fill=teal_fill, font=white_font, align=center)

    kpis = [
        ("تحقيق الإيرادات مقابل الهدف | Revenue vs target",
         '=IF(B9="","",TEXT(B9/B4,"0%"))',
         '=IF(B4<=0,0,IF(B9/B4>=115%,40,IF(B9/B4>=100%,35,IF(B9/B4>=80%,25,IF(B9/B4>=60%,15,IF(B9/B4>=50%,10,5))))))', 40),
        ("الاجتماعات الفعالة مقابل الهدف | Meetings vs target",
         '=IF(B10="","",TEXT(B10/B5,"0%"))',
         '=IF(B5<=0,0,IF(B10/B5>=120%,20,IF(B10/B5>=100%,15,IF(B10/B5>=80%,10,IF(B10/B5>=65%,5,0)))))', 20),
        ("نسبة إغلاق الصفقات | Deal closure rate",
         '=IF(OR(B10="",B10=0),"",TEXT(B11/B10,"0%"))',
         '=IF(B10=0,0,IF(B11/B10>30%,20,IF(B11/B10>=20%,15,IF(B11/B10>=10%,10,5))))', 20),
        ("كفاءة دورة المبيعات | Sales cycle efficiency",
         '=IF(OR(B12="",B6<=0),"",TEXT(B12/B6,"0%")&" of target")',
         '=IF(OR(B12="",B6<=0),0,IF(B12<B6,15,IF(B12=B6,10,IF(B12<=B6*1.2,8,3))))', 15),
        ("استخدام CRM وإدارة البيانات | CRM usage",
         '=IF(B13="","",TEXT(B13,"0%"))',
         '=IF(B13="",0,IF(B13>=100%,10,IF(B13>=75%,8,IF(B13>=60%,4,0))))', 10),
        ("رضا العملاء | Customer satisfaction",
         '=IF(B14="","",""&B14&" / 10")',
         '=IF(B14="",0,IF(B14>=9,10,IF(B14>=8,8,IF(B14>=7,5,3))))', 10),
    ]
    for i, (name, tier_f, pts_f, mx) in enumerate(kpis):
        r = 18 + i
        put(ws, f"A{r}", name, font=label_font, align=wrap)
        put(ws, f"B{r}", None, calc_fill, align=center)
        put(ws, f"C{r}", tier_f, calc_fill, align=center)
        put(ws, f"D{r}", pts_f, calc_fill, font=Font(bold=True), align=center)
        put(ws, f"E{r}", mx, align=center)

    put(ws, "A24", "المجموع الكلي | Total", font=Font(bold=True, size=12))
    put(ws, "D24", "=SUM(D18:D23)", teal_fill, Font(bold=True, size=14, color=WHITE), align=center)
    put(ws, "E24", 105, align=center)
    put(ws, "A25", "التقييم | Rating", font=Font(bold=True, size=12))
    rating_formula = ('=IF(D24>85,"ممتاز | Excellent",IF(D24>=75,"جيد جداً | Very Good",'
                      'IF(D24>=60,"يحتاج تحسين | Needs Improvement","مراجعة | Under Review")))')
    rc = put(ws, "C25", rating_formula, font=Font(bold=True, size=12), align=center)
    ws.merge_cells("C25:D25")

    green = PatternFill("solid", fgColor="DCFCE7")
    blue = PatternFill("solid", fgColor="DBEAFE")
    amber = PatternFill("solid", fgColor="FEF3C7")
    red = PatternFill("solid", fgColor="FEE2E2")
    ws.conditional_formatting.add("C25", FormulaRule(formula=['$D$24>85'], fill=green))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['AND($D$24>=75,$D$24<=85)'], fill=blue))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['AND($D$24>=60,$D$24<75)'], fill=amber))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['$D$24<60'], fill=red))

    ws.column_dimensions["A"].width = 46
    for col in "BCDE":
        ws.column_dimensions[col].width = 16

ws_r = wb.create_sheet("Sales Rep")
build_rep_sheet(ws_r)

def build_leader_sheet(ws):
    style_title(ws, "A1:E1", "بطاقة تقييم قائد فريق المبيعات | Team Leader Scorecard")
    put(ws, "A3", "الأهداف الشهرية | Monthly Targets", font=section_font)
    put(ws, "A4", "هدف مبيعات الفريق | Team sales target", font=label_font, align=wrap)
    put(ws, "B4", None, input_fill)
    put(ws, "A5", "هدف اجتماعات القائد | Leader meetings target", font=label_font, align=wrap)
    put(ws, "B5", None, input_fill)

    put(ws, "A7", "المنجز الفعلي | Actuals", font=section_font)
    actuals = [
        ("مبيعات الفريق المحققة | Team sales achieved", "B8", "#,##0", None),
        ("اجتماعات قادها/دعمها | Meetings led/supported", "B9", "0", None),
        ("إجمالي اجتماعات الفريق | Team total meetings", "B10", "0", None),
        ("صفقات الفريق المغلقة | Team deals closed", "B11", "0", None),
        ("أعضاء يحققون أهدافهم (عدد) | Members achieving goals", "B12", "0", None),
        ("إجمالي الأعضاء | Total members", "B13", "0", None),
    ]
    for i, (lbl, cell, fmt, _) in enumerate(actuals):
        put(ws, f"A{8+i}", lbl, font=label_font, align=wrap)
        put(ws, cell, None, input_fill, fmt=fmt)

    crm_list = '"100% دقة ووقت | Accurate,بعض أخطاء | Some errors,تكرار أخطاء | Repeated,إهمال | Neglect"'
    init_list = '"ناجحة | Success,غير ناجحة | Failed,خارجية | External,دون تطبيق | Proposed only,لا شيء | None"'
    dv_crm = DataValidation(type="list", formula1=crm_list, allow_blank=True)
    dv_init = DataValidation(type="list", formula1=init_list, allow_blank=True)
    ws.add_data_validation(dv_crm); ws.add_data_validation(dv_init)

    put(ws, "A14", "دقة CRM | CRM accuracy status", font=label_font, align=wrap)
    put(ws, "B14", None, input_fill)
    dv_crm.add("B14")
    put(ws, "A15", "حالة مبادرات التحسين | Initiative status", font=label_font, align=wrap)
    put(ws, "B15", None, input_fill)
    dv_init.add("B15")

    headers = ["المؤشر | KPI", "المنجز | Actual", "الفئة | Tier", "النقاط | Points", "الحد | Max"]
    for j, h in enumerate(headers):
        put(ws, f"{get_column_letter(j+1)}17", h, fill=teal_fill, font=white_font, align=center)

    kpis = [
        ("تحقيق أهداف الفريق | Team goal achievement",
         '=IF(B8="","",TEXT(B8/B4,"0%"))',
         '=IF(B4<=0,0,IF(B8/B4>=120%,30,IF(B8/B4>=100%,25,IF(B8/B4>=80%,20,IF(B8/B4>=50%,15,5)))))', 30),
        ("الإرشاد على الاجتماعات | Meetings organized/supported",
         '=IF(B9="","",TEXT(B9/B5,"0%"))',
         '=IF(B5<=0,0,IF(B9/B5>=120%,20,IF(B9/B5>=100%,15,IF(B9/B5>=80%,10,5))))', 20),
        ("تطوير أعضاء الفريق | Members achieving goals",
         '=IF(OR(B13="",B13=0),"",TEXT(B12/B13,"0%"))',
         '=IF(B13=0,0,IF(B12/B13>=90%,20,IF(B12/B13>=75%,15,IF(B12/B13>=60%,10,5))))', 20),
        ("نسبة إغلاق صفقات الفريق | Team closure rate",
         '=IF(OR(B10="",B10=0),"",TEXT(B11/B10,"0%"))',
         '=IF(B10=0,0,IF(B11/B10>50%,15,IF(B11/B10>=40%,12,IF(B11/B10>=30%,8,5))))', 15),
        ("أدوات الإدارة والتقارير | CRM & reporting tools",
         '=IF(B14="","","—")',
         '=IF(B14="100% دقة ووقت | Accurate",15,IF(B14="بعض أخطاء | Some errors",12,IF(B14="تكرار أخطاء | Repeated",9,IF(B14="إهمال | Neglect",2,0))))', 15),
        ("مبادرات تحسين الأداء | Performance initiatives",
         '=IF(B15="","","—")',
         '=IF(B15="ناجحة | Success",10,IF(B15="غير ناجحة | Failed",7,IF(B15="خارجية | External",5,IF(B15="دون تطبيق | Proposed only",3,IF(B15="لا شيء | None",0,0)))))', 10),
    ]
    for i, (name, tier_f, pts_f, mx) in enumerate(kpis):
        r = 18 + i
        put(ws, f"A{r}", name, font=label_font, align=wrap)
        put(ws, f"B{r}", None, calc_fill, align=center)
        put(ws, f"C{r}", tier_f, calc_fill, align=center)
        put(ws, f"D{r}", pts_f, calc_fill, font=Font(bold=True), align=center)
        put(ws, f"E{r}", mx, align=center)

    put(ws, "A24", "المجموع الكلي | Total", font=Font(bold=True, size=12))
    put(ws, "D24", "=SUM(D18:D23)", teal_fill, Font(bold=True, size=14, color=WHITE), align=center)
    put(ws, "E24", 110, align=center)
    put(ws, "A25", "التقييم | Rating", font=Font(bold=True, size=12))
    rating_formula = ('=IF(D24>88,"ممتاز | Excellent",IF(D24>=75,"جيد جداً | Very Good",'
                      'IF(D24>=65,"يحتاج تحسين | Needs Improvement","مراجعة | Under Review")))')
    put(ws, "C25", rating_formula, font=Font(bold=True, size=12), align=center)
    ws.merge_cells("C25:D25")

    green = PatternFill("solid", fgColor="DCFCE7")
    blue = PatternFill("solid", fgColor="DBEAFE")
    amber = PatternFill("solid", fgColor="FEF3C7")
    red = PatternFill("solid", fgColor="FEE2E2")
    ws.conditional_formatting.add("C25", FormulaRule(formula=['$D$24>88'], fill=green))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['AND($D$24>=75,$D$24<=88)'], fill=blue))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['AND($D$24>=65,$D$24<75)'], fill=amber))
    ws.conditional_formatting.add("C25", FormulaRule(formula=['$D$24<65'], fill=red))

    ws.column_dimensions["A"].width = 46
    for col in "BCDE":
        ws.column_dimensions[col].width = 18

ws_l = wb.create_sheet("Team Leader")
build_leader_sheet(ws_l)

wb.save(r"D:\HR\Majaara\KPI-Dashboard\KPI_Scorecard.xlsx")
print("Excel saved")
